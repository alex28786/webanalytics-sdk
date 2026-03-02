"""
CJA Configure — XLSX Storage Helpers
=====================================
Read/write dimension & metric definitions and sync state
from/to the Excel workbook.  Includes automatic backup
before writes and retry for PermissionError (file open in Excel).
"""

import shutil
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import XLSX_PATH, DATA_DIR

# ── Sheets ─────────────────────────────────────────────────────────────────
SHEET_DIMENSIONS = "dimensions"
SHEET_METRICS = "metrics"
SHEET_SYNC_LOG = "sync_log"
SHEET_CONFIG = "config"

# Column order for dimensions sheet
DIM_COLUMNS = [
    "name", "description", "id", "source", "scope",
    "expiration",
    "status", "component_id_dev", "component_id_prod",
    "last_synced_dev", "last_synced_prod",
]

# Column order for metrics sheet
MET_COLUMNS = [
    "name", "description", "id", "source", "scope",
    "deduplication",
    "status", "component_id_dev", "component_id_prod",
    "last_synced_dev", "last_synced_prod",
]

# Sync log columns
LOG_COLUMNS = [
    "timestamp", "env", "type", "items_total",
    "items_synced", "items_failed", "errors",
]


# ── Helpers ────────────────────────────────────────────────────────────────

def _ensure_data_dir():
    """Create the data directory if it doesn't exist."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _backup(path: Path):
    """Create a timestamped backup of the XLSX file."""
    if path.exists():
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
        shutil.copy2(path, backup)
        print(f"  ↳ Backup: {backup.name}")


def _save_with_retry(wb: Workbook, path: Path, retries: int = 3, delay: float = 2.0):
    """
    Save workbook with retry logic for PermissionError
    (file open in Excel).
    """
    for attempt in range(retries):
        try:
            wb.save(path)
            return
        except PermissionError:
            if attempt < retries - 1:
                print(
                    f"  ⚠ File is open — retrying in {delay}s "
                    f"(attempt {attempt + 1}/{retries})..."
                )
                time.sleep(delay)
            else:
                raise PermissionError(
                    f"Cannot write to {path.name}. "
                    "Please close it in Excel and try again."
                )


# ── Initialize workbook ───────────────────────────────────────────────────

def init_workbook(path: Path | None = None) -> Path:
    """
    Create a new XLSX with the expected sheets and column headers.
    Returns the path to the created file.
    """
    path = path or XLSX_PATH
    _ensure_data_dir()

    wb = Workbook()

    # Dimensions sheet (default sheet renamed)
    ws_dim = wb.active
    ws_dim.title = SHEET_DIMENSIONS
    ws_dim.append(DIM_COLUMNS)

    # Metrics sheet
    ws_met = wb.create_sheet(SHEET_METRICS)
    ws_met.append(MET_COLUMNS)

    # Sync log sheet
    ws_log = wb.create_sheet(SHEET_SYNC_LOG)
    ws_log.append(LOG_COLUMNS)

    # Config sheet
    ws_cfg = wb.create_sheet(SHEET_CONFIG)
    ws_cfg.append(["key", "value"])

    _save_with_retry(wb, path)
    print(f"  ✓ Initialized workbook: {path.name}")
    return path


# ── Read ───────────────────────────────────────────────────────────────────

def _read_sheet(sheet_name: str, path: Path | None = None) -> list[dict]:
    """Read a sheet into a list of dicts keyed by the header row."""
    path = path or XLSX_PATH
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        entry = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            entry[key] = val
        result.append(entry)

    return result


def read_dimensions(path: Path | None = None) -> list[dict]:
    """Read the dimensions sheet."""
    return _read_sheet(SHEET_DIMENSIONS, path)


def read_metrics(path: Path | None = None) -> list[dict]:
    """Read the metrics sheet."""
    return _read_sheet(SHEET_METRICS, path)


def read_sync_log(path: Path | None = None) -> list[dict]:
    """Read the sync_log sheet."""
    return _read_sheet(SHEET_SYNC_LOG, path)


# ── Write definitions ─────────────────────────────────────────────────────

def write_definitions(
    definitions: list[dict],
    sheet_name: str,
    columns: list[str],
    path: Path | None = None,
):
    """
    Write a list of definition dicts to a specific sheet.
    Overwrites all rows below the header.  Creates the workbook if needed.
    """
    path = path or XLSX_PATH
    _ensure_data_dir()

    if not path.exists():
        init_workbook(path)

    _backup(path)

    wb = load_workbook(path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
    else:
        ws = wb[sheet_name]
        # Clear all data rows, keep header
        ws.delete_rows(2, ws.max_row)

    for defn in definitions:
        row = [defn.get(col) for col in columns]
        ws.append(row)

    _save_with_retry(wb, path)
    print(f"  ✓ Wrote {len(definitions)} rows to [{sheet_name}]")


def write_dimensions(definitions: list[dict], path: Path | None = None):
    """Write dimension definitions to the dimensions sheet."""
    write_definitions(definitions, SHEET_DIMENSIONS, DIM_COLUMNS, path)


def write_metrics(definitions: list[dict], path: Path | None = None):
    """Write metric definitions to the metrics sheet."""
    write_definitions(definitions, SHEET_METRICS, MET_COLUMNS, path)


# ── Update sync state ─────────────────────────────────────────────────────

def update_sync_state(
    sheet_name: str,
    columns: list[str],
    updates: list[dict],
    env_name: str,
    path: Path | None = None,
):
    """
    After a sync, update component_id and last_synced columns
    for matched definitions.

    Parameters
    ----------
    sheet_name : str
    columns : list[str]  — column order
    updates : list[dict] — each has 'name', 'component_id'
    env_name : str       — 'dev' or 'prod'
    path : Path | None
    """
    path = path or XLSX_PATH
    if not path.exists():
        print("  ⚠ No workbook found; skipping state update.")
        return

    wb = load_workbook(path)
    ws = wb[sheet_name]

    # Build column index map from header row
    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    comp_col = f"component_id_{env_name}"
    ts_col = f"last_synced_{env_name}"
    status_col = "status"

    # Build update lookup
    update_map = {u["name"]: u["component_id"] for u in updates if u.get("component_id")}
    now = datetime.now().isoformat(timespec="seconds")

    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=col_idx.get("name", 1)).value
        if name in update_map:
            if comp_col in col_idx:
                ws.cell(row=row_idx, column=col_idx[comp_col]).value = update_map[name]
            if ts_col in col_idx:
                ws.cell(row=row_idx, column=col_idx[ts_col]).value = now
            if status_col in col_idx:
                ws.cell(row=row_idx, column=col_idx[status_col]).value = "Synced"

    _save_with_retry(wb, path)
    print(f"  ✓ Updated sync state for {len(update_map)} items in [{sheet_name}]")


# ── Sync log ───────────────────────────────────────────────────────────────

def append_sync_log(
    env: str,
    sync_type: str,
    items_total: int,
    items_synced: int,
    items_failed: int,
    errors: str = "",
    path: Path | None = None,
):
    """Append a row to the sync_log sheet."""
    path = path or XLSX_PATH
    _ensure_data_dir()

    if not path.exists():
        init_workbook(path)

    wb = load_workbook(path)

    if SHEET_SYNC_LOG not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SYNC_LOG)
        ws.append(LOG_COLUMNS)
    else:
        ws = wb[SHEET_SYNC_LOG]

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        env,
        sync_type,
        items_total,
        items_synced,
        items_failed,
        errors,
    ])

    _save_with_retry(wb, path)
    print(f"  ✓ Sync log entry added")
