"""
Tests — Storage (XLSX) Read/Write
==================================
Tests round-trip XLSX operations, backup creation, and sync log.
"""

import pytest
import sys
import os
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from storage import (
    init_workbook,
    write_dimensions,
    write_metrics,
    read_dimensions,
    read_metrics,
    update_sync_state,
    append_sync_log,
    read_sync_log,
    DIM_COLUMNS,
    MET_COLUMNS,
    SHEET_DIMENSIONS,
    SHEET_METRICS,
)


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Provide a temporary XLSX path."""
    return tmp_path / "test-workbook.xlsx"


class TestWorkbookInit:
    def test_creates_file(self, tmp_xlsx):
        path = init_workbook(tmp_xlsx)
        assert path.exists()

    def test_has_expected_sheets(self, tmp_xlsx):
        from openpyxl import load_workbook
        init_workbook(tmp_xlsx)
        wb = load_workbook(tmp_xlsx, read_only=True)
        assert "dimensions" in wb.sheetnames
        assert "metrics" in wb.sheetnames
        assert "sync_log" in wb.sheetnames
        assert "config" in wb.sheetnames
        wb.close()


class TestDimensionReadWrite:
    def test_round_trip(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        sample = [
            {
                "name": "Test Dim [v1]",
                "description": "Test dimension",
                "id": "variables/_experience.analytics.customDimensions.eVars.eVar1",
                "source": "_experience.analytics.customDimensions.eVars.eVar1",
                "scope": "hit",
                "expiration": None,
                "status": None,
                "component_id_dev": None,
                "component_id_prod": None,
                "last_synced_dev": None,
                "last_synced_prod": None,
            }
        ]
        write_dimensions(sample, tmp_xlsx)
        result = read_dimensions(tmp_xlsx)

        assert len(result) == 1
        assert result[0]["name"] == "Test Dim [v1]"
        assert result[0]["scope"] == "hit"

    def test_multiple_rows(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        rows = [
            {col: f"val_{col}_{i}" for col in DIM_COLUMNS}
            for i in range(10)
        ]
        write_dimensions(rows, tmp_xlsx)
        result = read_dimensions(tmp_xlsx)
        assert len(result) == 10

    def test_preserves_empty_on_no_file(self, tmp_xlsx):
        """Reading from non-existent file returns empty list."""
        result = read_dimensions(tmp_xlsx)
        assert result == []


class TestSyncStateUpdate:
    def test_updates_component_id(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        sample = [
            {
                "name": "Test Dim",
                "description": "test",
                "id": "test_id",
                "source": "test_source",
                "scope": "hit",
                "expiration": None,
                "status": "Pending",
                "component_id_dev": None,
                "component_id_prod": None,
                "last_synced_dev": None,
                "last_synced_prod": None,
            }
        ]
        write_dimensions(sample, tmp_xlsx)

        updates = [{"name": "Test Dim", "component_id": "comp_abc123"}]
        update_sync_state(SHEET_DIMENSIONS, DIM_COLUMNS, updates, "dev", tmp_xlsx)

        result = read_dimensions(tmp_xlsx)
        assert result[0]["component_id_dev"] == "comp_abc123"
        assert result[0]["status"] == "Synced"
        assert result[0]["last_synced_dev"] is not None


class TestSyncLog:
    def test_append_log(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        append_sync_log(
            env="dev",
            sync_type="dimensions",
            items_total=10,
            items_synced=8,
            items_failed=2,
            errors="test error",
            path=tmp_xlsx,
        )

        logs = read_sync_log(tmp_xlsx)
        assert len(logs) == 1
        assert logs[0]["env"] == "dev"
        assert logs[0]["items_synced"] == 8

    def test_multiple_log_entries(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        for i in range(3):
            append_sync_log(
                env="dev",
                sync_type=f"test_{i}",
                items_total=i,
                items_synced=i,
                items_failed=0,
                path=tmp_xlsx,
            )

        logs = read_sync_log(tmp_xlsx)
        assert len(logs) == 3


class TestBackup:
    def test_backup_created_on_write(self, tmp_xlsx):
        init_workbook(tmp_xlsx)
        sample = [
            {col: "test" for col in DIM_COLUMNS}
        ]
        write_dimensions(sample, tmp_xlsx)

        # Check that a backup was created
        backups = list(tmp_xlsx.parent.glob("*_backup_*"))
        assert len(backups) >= 1
